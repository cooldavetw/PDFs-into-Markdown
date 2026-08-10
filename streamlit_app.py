import streamlit as st
from docling.document_converter import DocumentConverter
import tempfile
import os
import logging
import io
import zipfile
import shutil
import subprocess
from pathlib import Path, PurePosixPath

from openpyxl import load_workbook

# Configure logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

DIRECT_CONVERSION_EXTENSIONS = ('.pdf', '.docx')
SPREADSHEET_EXTENSIONS = ('.xlsx', '.xls')
SUPPORTED_DOCUMENT_EXTENSIONS = DIRECT_CONVERSION_EXTENSIONS + ('.doc',) + SPREADSHEET_EXTENSIONS


def convert_document_to_markdown(document_path):
    result = st.session_state.converter.convert(document_path)
    return result.document.export_to_markdown()


def convert_doc_to_docx(doc_path):
    office_command = shutil.which('soffice') or shutil.which('libreoffice')
    if office_command is None:
        raise RuntimeError("需要 LibreOffice 才能轉換舊版 .doc 檔案。")

    with tempfile.TemporaryDirectory() as output_dir, tempfile.TemporaryDirectory() as profile_dir:
        libreoffice_env = os.environ.copy()
        libreoffice_env["HOME"] = profile_dir

        completed_process = subprocess.run(
            [
                office_command,
                f"-env:UserInstallation={Path(profile_dir).as_uri()}",
                '--headless',
                '--convert-to',
                'docx',
                '--outdir',
                output_dir,
                doc_path,
            ],
            capture_output=True,
            text=True,
            timeout=120,
            check=False,
            env=libreoffice_env,
        )

        output_path = os.path.join(
            output_dir,
            os.path.splitext(os.path.basename(doc_path))[0] + '.docx',
        )
        if completed_process.returncode != 0 or not os.path.exists(output_path):
            error_output = completed_process.stderr or completed_process.stdout
            raise RuntimeError(f"轉換 DOC 檔案時發生錯誤：{error_output.strip()}")

        with tempfile.NamedTemporaryFile(delete=False, suffix='.docx') as tmp_file:
            with open(output_path, 'rb') as converted_file:
                shutil.copyfileobj(converted_file, tmp_file)
            return tmp_file.name


def convert_xls_to_xlsx(xls_path):
    office_command = shutil.which('soffice') or shutil.which('libreoffice')
    if office_command is None:
        raise RuntimeError("需要 LibreOffice 才能轉換舊版 .xls 檔案。")

    with tempfile.TemporaryDirectory() as output_dir, tempfile.TemporaryDirectory() as profile_dir:
        libreoffice_env = os.environ.copy()
        libreoffice_env["HOME"] = profile_dir

        completed_process = subprocess.run(
            [
                office_command,
                f"-env:UserInstallation={Path(profile_dir).as_uri()}",
                '--headless',
                '--convert-to',
                'xlsx',
                '--outdir',
                output_dir,
                xls_path,
            ],
            capture_output=True,
            text=True,
            timeout=120,
            check=False,
            env=libreoffice_env,
        )

        output_path = os.path.join(
            output_dir,
            os.path.splitext(os.path.basename(xls_path))[0] + '.xlsx',
        )
        if completed_process.returncode != 0 or not os.path.exists(output_path):
            error_output = completed_process.stderr or completed_process.stdout
            raise RuntimeError(f"轉換 XLS 檔案時發生錯誤：{error_output.strip()}")

        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            with open(output_path, 'rb') as converted_file:
                shutil.copyfileobj(converted_file, tmp_file)
            return tmp_file.name


def stringify_spreadsheet_value(value):
    if value is None:
        return ""

    text = str(value)
    return text.replace('|', '\\|').replace('\r\n', '<br>').replace('\n', '<br>').replace('\r', '<br>')


def trim_empty_spreadsheet_edges(rows):
    while rows and all(value in (None, "") for value in rows[-1]):
        rows.pop()

    if not rows:
        return []

    last_column_index = -1
    for row in rows:
        for index, value in enumerate(row):
            if value not in (None, ""):
                last_column_index = max(last_column_index, index)

    return [row[:last_column_index + 1] for row in rows]


def worksheet_to_markdown_table(worksheet):
    rows = [
        list(row)
        for row in worksheet.iter_rows(values_only=True)
    ]
    rows = trim_empty_spreadsheet_edges(rows)

    if not rows:
        return "_找不到資料。_"

    column_count = max(len(row) for row in rows)
    normalized_rows = [
        row + [None] * (column_count - len(row))
        for row in rows
    ]
    header = [
        stringify_spreadsheet_value(value) or f"欄位 {index + 1}"
        for index, value in enumerate(normalized_rows[0])
    ]
    body_rows = normalized_rows[1:]

    markdown_lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join(["---"] * column_count) + " |",
    ]
    for row in body_rows:
        markdown_lines.append(
            "| " + " | ".join(stringify_spreadsheet_value(value) for value in row) + " |"
        )

    return "\n".join(markdown_lines)


def convert_spreadsheet_to_markdown(spreadsheet_path):
    workbook = load_workbook(spreadsheet_path, data_only=True, read_only=True)
    markdown_sections = []

    try:
        for worksheet in workbook.worksheets:
            markdown_sections.append(
                f"## {worksheet.title}\n\n{worksheet_to_markdown_table(worksheet)}"
            )
    finally:
        workbook.close()

    return "\n\n".join(markdown_sections)


def convert_path_to_markdown(file_path, file_extension):
    converted_docx_path = None
    converted_xlsx_path = None

    try:
        conversion_path = file_path
        if file_extension == '.doc':
            converted_docx_path = convert_doc_to_docx(file_path)
            conversion_path = converted_docx_path
        elif file_extension == '.xls':
            converted_xlsx_path = convert_xls_to_xlsx(file_path)
            conversion_path = converted_xlsx_path

        if file_extension in SPREADSHEET_EXTENSIONS:
            return convert_spreadsheet_to_markdown(conversion_path)

        return convert_document_to_markdown(conversion_path)
    finally:
        if converted_docx_path and os.path.exists(converted_docx_path):
            os.unlink(converted_docx_path)
            logger.debug("Temporary converted DOCX file deleted")
        if converted_xlsx_path and os.path.exists(converted_xlsx_path):
            os.unlink(converted_xlsx_path)
            logger.debug("Temporary converted XLSX file deleted")


def convert_uploaded_document(uploaded_file):
    file_extension = os.path.splitext(uploaded_file.name)[1].lower()

    with tempfile.NamedTemporaryFile(delete=False, suffix=file_extension) as tmp_file:
        tmp_file.write(uploaded_file.getvalue())
        tmp_path = tmp_file.name
        logger.debug(f"Temporary file created at: {tmp_path}")

    try:
        return convert_path_to_markdown(tmp_path, file_extension)
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)
            logger.debug("Temporary file deleted")


def is_supported_document(filename):
    return filename.lower().endswith(SUPPORTED_DOCUMENT_EXTENSIONS)


def markdown_name_for_zip_member(member_name, used_names):
    path = PurePosixPath(member_name.replace('\\', '/'))
    safe_parts = [
        part for part in path.parts
        if part not in ('', '.', '..') and not part.startswith('/')
    ]
    path = PurePosixPath(*safe_parts) if safe_parts else PurePosixPath('document.pdf')
    output_name = str(path.with_suffix('.md'))

    if output_name not in used_names:
        used_names.add(output_name)
        return output_name

    stem = str(path.with_suffix(''))
    counter = 2
    while True:
        candidate = f"{stem}-{counter}.md"
        if candidate not in used_names:
            used_names.add(candidate)
            return candidate
        counter += 1


def convert_uploaded_zip(uploaded_file):
    markdown_zip_buffer = io.BytesIO()
    converted_count = 0
    used_names = set()
    markdown_files = []
    uploaded_document_names = []

    with zipfile.ZipFile(io.BytesIO(uploaded_file.getvalue())) as input_zip:
        document_members = [
            info for info in input_zip.infolist()
            if not info.is_dir() and is_supported_document(info.filename)
        ]

        if not document_members:
            raise ValueError("ZIP 檔案中不包含任何支援的文件檔案。")

        with zipfile.ZipFile(markdown_zip_buffer, mode='w', compression=zipfile.ZIP_DEFLATED) as output_zip:
            for member in document_members:
                file_extension = os.path.splitext(member.filename)[1].lower()
                uploaded_document_names.append(member.filename)

                with tempfile.NamedTemporaryFile(delete=False, suffix=file_extension) as tmp_file:
                    tmp_file.write(input_zip.read(member))
                    tmp_path = tmp_file.name
                    logger.debug(f"Temporary ZIP document created at: {tmp_path}")

                try:
                    markdown_text = convert_path_to_markdown(tmp_path, file_extension)
                    output_name = markdown_name_for_zip_member(member.filename, used_names)
                    output_zip.writestr(output_name, markdown_text)
                    markdown_files.append((output_name, markdown_text))
                    converted_count += 1
                finally:
                    if os.path.exists(tmp_path):
                        os.unlink(tmp_path)
                        logger.debug("Temporary ZIP document deleted")

    markdown_zip_buffer.seek(0)
    return markdown_zip_buffer.getvalue(), converted_count, markdown_files, uploaded_document_names


def display_uploaded_documents(document_names, key):
    st.subheader("已上傳的文件")
    st.text_area(
        "已上傳的文件",
        "\n".join(document_names),
        height=min(300, 70 + (len(document_names) * 24)),
        key=key,
        label_visibility="collapsed",
    )


def display_markdown_preview(markdown_text, key):
    st.subheader("轉換後的 Markdown")
    st.text_area(
        "轉換後的 Markdown",
        markdown_text,
        height=500,
        key=key,
        label_visibility="collapsed",
    )


def display_zip_markdown_previews(markdown_files):
    st.subheader("轉換後的 Markdown 檔案")
    for index, (filename, markdown_text) in enumerate(markdown_files):
        with st.expander(filename):
            st.text_area(
                "轉換後的 Markdown",
                markdown_text,
                height=400,
                key=f"zip_markdown_preview_{index}",
                label_visibility="collapsed",
            )


st.set_page_config(page_title="文件轉 Markdown 轉換器")

# Custom CSS for better layout
st.markdown("""
    <style>    
        .stFileUploader {
            padding: 1rem;
        }
        
        button[data-testid="stFileUploaderButtonPrimary"] {
            background-color: #000660 !important;
            border: none !important;
            color: white !important;
        }

        .stButton button {
            background-color: #006666;
            border: none !important;
            color: white;
            padding: 0.5rem 2rem !important;
        }
        .stButton button:hover {
            background-color: #008080 !important;
            color: white !important;
            border-color: #008080 !important;
        }
        .upload-text {
            font-size: 1.2rem;
            margin-bottom: 1rem;
        }
        div[data-testid="stFileUploadDropzone"]:hover {
            border-color: #006666 !important;
            background-color: rgba(0, 102, 102, 0.05) !important;
        }
    </style>
""", unsafe_allow_html=True)

st.title("文件 (PDF, Word, Excel) Markdown 轉換器")

# Initialize session state if it doesn't exist
if 'converter' not in st.session_state:
    try:
        st.session_state.converter = DocumentConverter()
        logger.debug("Converter successfully created")
    except Exception as e:
        logger.error(f"Error creating converter: {str(e)}")
        st.error(f"建立轉換器時發生錯誤：{str(e)}")
        st.stop()

# Main upload area
uploaded_file = st.file_uploader(
    "上傳 PDF、Word 文件、Excel 活頁簿，或包含文件的 ZIP 檔案",
    type=['pdf', 'doc', 'docx', 'xls', 'xlsx', 'zip'],
    key='pdf_uploader',
    help="拖放或點選一個 PDF、Word、Excel 檔案，或一個包含文件的 ZIP 檔案（最大 200MB）"
)

# Unified convert button
convert_clicked = st.button("轉換為 Markdown", type="primary")

# Process uploaded file
if convert_clicked:
    if uploaded_file is not None:
        file_extension = os.path.splitext(uploaded_file.name)[1].lower()

        if file_extension in SUPPORTED_DOCUMENT_EXTENSIONS:
            try:
                with st.spinner('正在轉換檔案...'):
                    markdown_text = convert_uploaded_document(uploaded_file)
                    output_filename = os.path.splitext(uploaded_file.name)[0] + '.md'

                    st.success("轉換完成！")
                    display_uploaded_documents([uploaded_file.name], "uploaded_document_list")
                    display_markdown_preview(markdown_text, "uploaded_document_markdown_preview")
                    st.download_button(
                        label="下載 Markdown 檔案",
                        data=markdown_text,
                        file_name=output_filename,
                        mime="text/markdown"
                    )

            except Exception as e:
                logger.error(f"Error processing document file: {str(e)}")
                st.error(f"處理文件檔案時發生錯誤：{str(e)}")

        elif file_extension == '.zip':
            try:
                with st.spinner('正在轉換 ZIP 中的文件檔案...'):
                    markdown_zip, converted_count, markdown_files, uploaded_document_names = convert_uploaded_zip(uploaded_file)
                    output_filename = os.path.splitext(uploaded_file.name)[0] + '_markdown.zip'

                    st.success(f"轉換完成！已轉換 {converted_count} 個文件檔案。")
                    display_uploaded_documents(uploaded_document_names, "zip_uploaded_document_list")
                    display_zip_markdown_previews(markdown_files)
                    st.download_button(
                        label="下載 Markdown ZIP 檔案",
                        data=markdown_zip,
                        file_name=output_filename,
                        mime="application/zip"
                    )

            except zipfile.BadZipFile:
                logger.error("Uploaded file is not a valid ZIP file")
                st.error("上傳的檔案不是有效的 ZIP 檔案。")
            except Exception as e:
                logger.error(f"Error processing ZIP file: {str(e)}")
                st.error(f"處理 ZIP 檔案時發生錯誤：{str(e)}")

        else:
            st.error("請上傳 PDF、Word 文件、Excel 活頁簿，或包含支援文件的 ZIP 檔案。")
    else:
        st.warning("請先上傳 PDF、Word 文件、Excel 活頁簿，或包含支援文件的 ZIP 檔案。")
